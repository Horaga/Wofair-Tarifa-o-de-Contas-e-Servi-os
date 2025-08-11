# app.py
# -*- coding: utf-8 -*-
import re
from decimal import Decimal, ROUND_HALF_UP
from dataclasses import dataclass
from typing import Dict, Optional, Tuple, List
import io
from openpyxl.utils import get_column_letter
import streamlit as st
import pandas as pd
import pdfplumber

# ============== DICIONÁRIO FIXO DE VALORES UNITÁRIOS ==============
UNITARIOS: Dict[str, float] = {
  "*SAQUE UL": 1.21,
  "BIOMETRIA SAQUE U": 1.10,
  "BIOM SQ CT DIGIT": 0.71,
  "BIOMETRIA SQ POU": 0.68,
  "*PAG INSS UL": 0.71,
  "*PAG INSS UL SOC": 0.71,
  "*SAQUE POUPANCA F": 0.68,
  "SAQUE EMERG S/CA": 0.78,
  "*SAQUE CT SOCIAL": 0.74,
  "*SQ CTA AUX BRASI": 0.78,
  "2-BENEF SOCIAIS": 0.89,
  "AGIBANK": 0.90,
  "*PAG FGTS": 1.28,
  "*DEPOSITO UL": 1.24,
  "2- CARTOES": 0.99,
  "*CAESB": 0.96,
  "*COPASA": 0.96,
  "*EMBASA": 0.96,
  "*SANEAGO": 0.96,
  "*C.AGUASLI": 0.96,
  "NEOENERGIA": 0.96,
  "CLIENT CO": 0.96,
  "OI": 0.96,
  "*VIVO SE": 0.96,
  "*TELEGOC": 0.96,
  "*TELEBC": 0.96,
  "*VIVO MG": 0.96,
  "*TELESP CL": 0.96,
  "*VIVOFIXO": 0.96,
  "*TIM": 0.96,
  "CLARO SA": 0.96,
  "*CLAROTV": 0.96,
  "*SKY": 0.96,
  "2-GPS S/ BARRA UL": 0.86,
  "*VIVO DF": 1.04,
  "*VIVO FIXO": 1.04,
  "*TIMCEL": 1.04,
  "*CLARODDDS": 1.04,
  "*DIRECTVSK": 1.04,
  "2-H AFINZ": 1.04,
  "SHOW DA FE": 1.04,
  "MUN CAB GRANDE": 0.96,
  "2-SEFAZ/GO-DARE": 0.96,
  "2-SEFAZ/DF": 0.96,
  "2-IPVA MG": 0.96,
  "2-DETRAN GO": 0.96,
  "GPS COM BARRA": 0.86,
  "2-SIMPLESNACION": 0.86,
  "DARF NUMERADO": 0.86,
  "DAE": 0.86,
  "2-DETRAN-DF": 0.96,
  "2-DER": 0.96,
  "*CLARO CO": 0.744,
  "*TCODFCE": 0.575,
  "*TIM-GSM": 0.51,
  "*VIVO PI": 0.54,
  "PEC COBCAIXA": 0.94,
  "* NPC BOLETOS CAI": 0.94,
  "* NPC BOLETOS OUT": 1.24,
  "*PIX SAQUE": 0.74,
  "*PIX SAQUE CAIXA": 0.74,
  "*SALDO": 0.21,
  "*ATIVA CAIXA TEM": 1.02
}

# ========================= UTIL =========================
def q2(x: Decimal) -> Decimal:
    return x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

# ========================= PARSER PDF =========================

SECTION_HEADERS = {"PAGAMENTOS", "RECEBIMENTOS", "NEGOCIAL", "SERVICOS", "SERVIÇOS"}

# Regex 1 (original, coluna a coluna)
ITEM_RE_STRICT = re.compile(
    r"""^(?P<tipo>[\*\wÀ-ÖØ-öø-ÿ0-9/\-\.\s]+?)   # Tipo
        [ \t]{2,}                                 # separador
        (?P<qtde>\d+)                             # Qtde
        [ \t]{2,}                                 # separador
        [\d\.,]+                                  # Valor (ignorado)
        \s*$""",
    re.VERBOSE
)

# Regex 2 (fallback): captura pelo final da linha: ... <qtde> <valor>
ITEM_RE_FALLBACK = re.compile(
    r"""^(?P<tipo>.+?)\s+ (?P<qtde>\d+)\s+ (?P<valor>[\d\.,]+)\s*$""",
    re.VERBOSE
)

def _normalize_spaces(s: str) -> str:
    # normaliza variações de espaço que vêm do PDF
    return (s.replace("\u00A0", " ")
             .replace("\u2007", " ")
             .replace("\u2002", " ")
             .replace("\u2003", " "))

@dataclass
class Item:
    secao: str
    tipo: str
    qtde: int

def extract_pdf_text(uploaded_file) -> str:
    with pdfplumber.open(uploaded_file) as pdf:
        pages = []
        for p in pdf.pages:
            txt = p.extract_text(x_tolerance=1, y_tolerance=1) or ""
            pages.append(txt)
        return "\n".join(pages)

def parse_items(text: str) -> Tuple[List[Item], Optional[str]]:
    # normaliza acentos em "SERVIÇOS" e espaços invisíveis
    text = _normalize_spaces(text).replace("SERVIÇOS", "SERVICOS")
    lines = [ln.rstrip() for ln in text.splitlines()]

    # data de referência (opcional)
    mdate = re.search(r"(\d{2}/[A-Z]{3}/\d{2})", text)
    data_ref = mdate.group(1) if mdate else None

    current = None
    items: List[Item] = []

    for raw_ln in lines:
        s = _normalize_spaces(raw_ln).strip()
        if not s:
            continue

        # detecta mudanças de seção
        if s in SECTION_HEADERS:
            current = s
            continue

        if not current:
            continue

        # ignora cabeçalho "TIPO----QTDE----VALOR" e fecha ao encontrar "TOTAL"
        if s.startswith("TIPO") or s.startswith("TOTAL"):
            if s.startswith("TOTAL"):
                current = None
            continue

        # tenta regex estrito
        m1 = ITEM_RE_STRICT.match(s)
        if m1:
            tipo = re.sub(r"\s{2,}", " ", m1.group("tipo")).strip()
            qtde = int(m1.group("qtde"))
            items.append(Item(secao=current, tipo=tipo, qtde=qtde))
            continue

        # tenta fallback
        m2 = ITEM_RE_FALLBACK.match(s)
        if m2:
            tipo = re.sub(r"\s{2,}", " ", m2.group("tipo")).strip()
            qtde = int(m2.group("qtde"))
            # confere que não é o cabeçalho
            if tipo.upper().startswith("TIPO"):
                continue
            items.append(Item(secao=current, tipo=tipo, qtde=qtde))
            continue

        # (opcional) debug: linhas que não casaram
        # st.write("Linha não casou:", repr(s))

    return items, data_ref

# ========================= UI =========================
st.set_page_config(page_title="Wofair• Loterias", page_icon="📄", layout="wide")
st.title("📄 Calculadora de Tarifação de Contas e Serviços")

uploaded = st.file_uploader("Envie o PDF do relatório diário", type=["pdf"])

if not uploaded:
    st.info("Envie o PDF para começar.")
    st.stop()

# Extrai texto
try:
    raw = extract_pdf_text(uploaded)
except Exception as e:
    st.error(f"Falha ao ler PDF: {e}")
    st.stop()

items, data_ref = parse_items(raw)

# Monta DF com chaves nomeadas (evita colunas 0/1/2)
df = pd.DataFrame(
    [{"Seção": it.secao, "Tipo": it.tipo, "Qtde": it.qtde} for it in items],
    columns=["Seção", "Tipo", "Qtde"]
)

# Guardas defensivos
df = pd.DataFrame(
    [{"Seção": it.secao, "Tipo": it.tipo, "Qtde": it.qtde} for it in items],
    columns=["Seção", "Tipo", "Qtde"]
)

if df.empty:
    st.error("Não consegui extrair itens (o DataFrame ficou vazio). É comum quando o PDF usa espaços não‑quebrantes. Já ativei a normalização; tente novamente.")
    st.caption("Se persistir, ative a linha de debug no parser para ver qual linha não casou.")
    st.stop()

if not {"Tipo", "Qtde"}.issubset(df.columns):
    st.error("Não consegui montar as colunas esperadas ('Tipo', 'Qtde').")
    st.write("Colunas encontradas:", df.columns.tolist())
    st.dataframe(df, use_container_width=True, hide_index=True)
    st.stop()

# Cálculo Valor Esperado usando UNITARIOS fixo
def valor_esperado(tipo: str, qtde: int) -> Optional[float]:
    vu = UNITARIOS.get(tipo)
    if vu is None:
        return None
    return float(q2(Decimal(qtde) * Decimal(str(vu))))

df["Valor Unitário"] = df["Tipo"].map(lambda t: float(UNITARIOS[t]) if t in UNITARIOS else None)
df["Valor Esperado"] = df.apply(lambda r: valor_esperado(r["Tipo"], r["Qtde"]), axis=1)

# Métricas
total_esperado = q2(sum(Decimal(str(v)) for v in df["Valor Esperado"].dropna())) if df["Valor Esperado"].notna().any() else Decimal("0.00")
faltando_unit = int(df["Valor Unitário"].isna().sum())

c1, c2 = st.columns(2)
fmt = lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
c1.metric("Total Esperado (todos os serviços)", fmt(total_esperado))
c2.metric("Tipos sem valor unitário definido", str(faltando_unit))

# Resumos
st.markdown("### Resumos")
colA, colB = st.columns(2)

resumo_tipo = (
    df.groupby("Tipo", as_index=False)
      .agg(Qtde=("Qtde", "sum"), Valor_Esperado=("Valor Esperado", "sum"))
      .sort_values(["Valor_Esperado", "Qtde"], ascending=[False, False])
)
colA.write("**Por Tipo**")
colA.dataframe(resumo_tipo, use_container_width=True, hide_index=True)

resumo_secao = (
    df.groupby("Seção", as_index=False)
      .agg(Qtde=("Qtde", "sum"), Valor_Esperado=("Valor Esperado", "sum"))
      .sort_values("Valor_Esperado", ascending=False)
)
colB.write("**Por Seção**")
colB.dataframe(resumo_secao, use_container_width=True, hide_index=True)

# Exportar
st.markdown("---")
st.subheader("⬇️ Exportar")
# --- Exportar para Excel (.xlsx) ---
buffer = io.BytesIO()
with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    # Abas
    df.to_excel(writer, index=False, sheet_name="Itens")
    resumo_tipo.to_excel(writer, index=False, sheet_name="Resumo por Tipo")
    resumo_secao.to_excel(writer, index=False, sheet_name="Resumo por Seção")

    # ===== formatação (aba Itens) =====
    wb = writer.book
    ws = writer.sheets["Itens"]

    # Largura de colunas (ajuste se quiser)
    widths = {
        1: 12,   # Seção
        2: 40,   # Tipo
        3: 10,   # Qtde
        4: 14,   # Valor Unitário
        5: 16,   # Valor Esperado
    }
    for col_idx, width in widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Formato de moeda nas colunas correspondentes (se existirem)
    for col_name in ["Valor Unitário", "Valor Esperado"]:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1  # 1-based
            col_letter = get_column_letter(col_idx)
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = 'R$ #,##0.00'

# Botão de download
st.download_button(
    label="Baixar Excel (.xlsx)",
    data=buffer.getvalue(),
    file_name="valores_esperados.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)
